import sqlite3
import os
import hashlib
from datetime import datetime
import sys
from openpyxl import Workbook
from openpyxl.styles import Font
import socket

DB_CONN = None


def get_base_file():
    """Get the same base filename"""
    return os.path.splitext(os.path.basename(__file__))[0]


def connect_db():
    """Connect or Create to a database"""
    global DB_CONN
    if DB_CONN is None:
        try:
            dbfile = get_base_file() + ".db"
            DB_CONN = sqlite3.connect(dbfile, timeout=2)
            print(f"Successfully connected SQLite instance: {dbfile}")
        except sqlite3.Error as e:
            print(f"Error creating SQLite instance: {e}")
            return None
    return DB_CONN


def close_db():
    global DB_CONN
    if DB_CONN:
        DB_CONN.close()
        DB_CONN = None


def query_database(query, args="", table="files"):
    """Query the database for changes"""
    conn = connect_db()
    if conn is None:
        print("Error: No database connection provided")
        return False
    try:
        cursor = conn.cursor()
        cursor.execute(query, args)
    except sqlite3.Error as e:
        print(f"Error querying master database: {e}")
        if cursor != None:
            cursor.close()
    finally:
        conn.commit()
        if cursor != None:
            cursor.close()


def fetch_database(query, args=""):
    """Fetch results to display"""
    conn = connect_db()
    result = 0
    if conn is None:
        print("Error: No database connection provided")
        return False
    try:
        cursor = conn.cursor()
        result = cursor.execute(query, args).fetchall()
    except sqlite3.Error as e:
        print(f"Error querying master database: {e}")
        if cursor != None:
            cursor.close()
    finally:
        if cursor != None:
            cursor.close()
    return result


def list_tables():
    """Show all tables"""
    query = "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
    tables = fetch_database(query)
    if len(tables) > 0:
        print("Tables in database:")
        print("-" * 40)
        for table in tables:
            table_name = table[0]
            # if not table_name.startswith("sqlite_"):
            print(f"  - {table_name}")

        # return [table[0] for table in tables if not table[0].startswith("sqlite_")]
    else:
        print("There are no tables")


def list_files(table="files"):
    """Show all files"""
    query = f"SELECT file, md5 from {table}"
    files = fetch_database(query)
    if len(files) > 0:
        print(f"{len(files)} Files in database:")
        print("-" * 40)
        for file in files:
            print(file)
    else:
        print("There are no files")


def check_table(table):
    """Checks if a SQLite DB Table exists"""
    query = "SELECT name FROM sqlite_master WHERE type='table' AND name=?"
    return len(fetch_database(query, (table,)))


def get_column_names(table_name="files"):
    """Check the columns for the table"""
    query = f"PRAGMA table_info({table_name})"
    columns = fetch_database(query)
    return [col[1] for col in columns]


def delete_table(table):
    """Deletes a SQLite DB Table"""
    query = f"DROP TABLE IF EXISTS {table}"
    query_database(query)


def create_hashtable(table="files"):
    """Creates a SQLite DB Table"""
    query = f"""
        CREATE TABLE IF NOT EXISTS {table} (
           file TEXT,
           md5 TEXT
    )"""
    query_database(query)


def create_hashtable_idx(table="files"):
    """Creates a SQLite DB Table Index"""
    query = f"""
    CREATE INDEX IF NOT EXISTS idx_{table} 
    ON {table} (file, md5)"""
    query_database(query)


def update_hashtable(file, md5, table="files"):
    """Update MD5 hash for a file"""
    query = f"UPDATE {table} SET md5=? WHERE file=?"
    args = (md5, file)
    if check_table(table):
        query_database(query, args)
    else:
        setup_hashtable(file, md5, table)


def insert_hashtable(file, md5, table="files"):
    """Insert file into the Files table"""
    query = f"INSERT INTO {table} (file, md5) VALUES (?, ?)"
    args = (file, md5)
    query_database(query, args)


def delete_rows(query, args, table="files"):
    # query = f"DELETE FROM {table} WHERE file LIKE ?"
    # args = ("%.xlsx",)
    query_database(query, args)


def setup_hashtable(file, md5, table="files"):
    """Setup the Files table"""
    create_hashtable(table)
    create_hashtable_idx(table)
    insert_hashtable(file, md5, table)


def md5indb(file, table="files"):
    """Check md5 hash exists"""
    query = f"SELECT md5 FROM {table} WHERE file=?"
    return fetch_database(query, (file,))


def haschanged(file, md5, table="files"):
    """Check if file has changed"""
    current_md5 = md5indb(file, table)
    if not current_md5:
        insert_hashtable(file, md5, table)
        return True
    elif current_md5[0][0] != md5:
        update_hashtable(file, md5, table)
        return True
    else:
        return False


def get_fileext(file):
    """Get the file extension"""
    return os.path.splitext(file)[1]


def get_moddate(file):
    """Get the file modified date"""
    try:
        mtime = os.path.getmtime(file)
        # return datetime.fromtimestamp(mtime).strftime("%Y-%m-%d %H:%M:%S")
    except OSError as e:
        print(f"Error getting modified date for {file}: {e}")
        mtime = 0
    return str(mtime)


def md5short(file):
    """Get md5 file hash tag"""
    return hashlib.md5(str(file + "|" + get_moddate(file)).encode("utf-8")).hexdigest()


def check_filechanges(folder, exclude, ws):
    changed = False
    for subdir, dirs, files in os.walk(folder):
        dirs[:] = [d for d in dirs if not d.startswith(".")]
        files[:] = [f for f in files if not f.startswith(".")]
        for file in files:
            origin = os.path.join(subdir, file)
            if os.path.isfile(origin):
                file_ext = get_fileext(file)
                if not file_ext in exclude:
                    md5 = md5short(file)

                    if haschanged(file, md5):
                        now = get_datetime("%d-%d-%Y %H:%M:%S")
                        dt = now.split(" ")
                        row_xlsreport(ws, file, origin, subdir, dt[0], dt[1])
                        print(f"{origin} has changed on {now}")
                        changed = True
    return changed


def load_folders():
    folders = []
    extensions = []
    config = get_base_file() + ".ini"
    if os.path.isfile(config):
        cfile = open(config, "r")
        for line in cfile.readlines():
            line = line.strip()
            if "|" in line:
                folder_path = os.path.abspath(line.split("|")[0])
                if folder_path not in folders:
                    folders.append(folder_path)
                exts = line.split("|")[1]
                if len(exts) > 0:
                    extl = exts.split(",")
                    extensions.append(extl)
            else:
                folder_path = os.path.abspath(line.split("|")[0])
                if folder_path not in folders:
                    folders.append(folder_path)
                extensions.append([])
        cfile.close()
    return folders, extensions


def run_filechanges(ws):
    changed = False
    folders_exts = load_folders()
    for i, folder in enumerate(folders_exts[0]):
        exts = folders_exts[1]
        changed = check_filechanges(folder, exts[i], ws)
    return changed


def execute(args):
    changes = 0
    if len(args) > 1:
        if args[1].lower() == "--loop":
            wb, ws, st = start_xlsreport()
            try:
                while True:
                    if run_filechanges(ws):
                        changes += 1
            except KeyboardInterrupt:
                print("Stopped...")
                if changes > 0:
                    print(changes)
    else:
        wb, ws, st = start_xlsreport()
        if run_filechanges(ws):
            end_xlsreport(wb, st)


def get_datetime(format):
    today = datetime.today()
    return today.strftime(format)


def start_xlsreport():
    wb = Workbook()
    ws = wb.active
    ws.title = socket.gethostname()
    st = get_datetime("%d-%b-%Y %H_%M_%S")
    header_xlsreport(ws)
    return wb, ws, st


def end_xlsreport(wb, st):
    dt = f' from {st} to {get_datetime("%d-%b-%Y %H_%M_%S")}'
    file = get_base_file() + dt + ".xlsx"
    wb.save(file)
    close_db()


def header_xlsreport(ws):
    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Full File Name")
    ws.cell(row=1, column=3, value="Folder Name")
    ws.cell(row=1, column=4, value="Date")
    ws.cell(row=1, column=5, value="Time")

    a1 = ws["A1"]
    b1 = ws["B1"]
    c1 = ws["C1"]
    d1 = ws["D1"]
    e1 = ws["E1"]

    a1.font = Font(color="000000", bold=True)
    b1.font = Font(color="000000", bold=True)
    c1.font = Font(color="000000", bold=True)
    d1.font = Font(color="000000", bold=True)
    e1.font = Font(color="000000", bold=True)


def get_lastrow(ws):
    row = 1
    for cell in ws["A"]:
        if cell.value is None:
            break
        else:
            row += 1
    return row


def row_xlsreport(ws, file, ffilename, folder, d, t):
    row = get_lastrow(ws)
    ws.cell(row=row, column=1, value=file)
    ws.cell(row=row, column=2, value=ffilename)
    ws.cell(row=row, column=3, value=folder)
    ws.cell(row=row, column=4, value=d)
    ws.cell(row=row, column=5, value=t)


if __name__ == "__main__":
    execute(sys.argv)
    # create_hashtable()
    # delete_table("files")
    list_tables()
    list_files()
